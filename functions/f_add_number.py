from openpyxl import load_workbook


async def add_number(user_id, nick, reg_time):
    try:
        file = "/Config.xlsx"
    except OSError:
        file = "\Config.xlsx"

    wb_obj = load_workbook(filename=file)
    wsheet = wb_obj['Contacts']
    for num, ID in wsheet.iter_rows(values_only=True):
        if ID == user_id:
            return True
    return False