import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os, datetime


async def make_chart_guntt(data, masters, dates):
    len_dates = len(dates)
    len_mas = len(masters)
    rows = []
    font = {'size': 5}
    plt.rc('font', **font)
    fig, gnt = plt.subplots(1, len_mas, sharex=True, sharey=True)
    for x in range(len_mas):
        gnt[x].set_title(masters[x])
    gnt[0].set_ylim(1, len_dates + 1)
    gnt[0].set_xlim(9, 18)
    plt.xticks([9, 10, 11, 12, 13, 14, 15, 16, 17, 18])
    for num in range(1, len_dates + 1):
        rows.append(num)
        dates[num-1] += '\n\n\n\n'
    gnt[0].set_yticks(rows)
    gnt[0].set_yticklabels(dates)
    for y in range(0, len_mas):
        z = 1
        gnt[y].grid(True)
        for master in data[y]:
            gnt[y].broken_barh(master, (z, 1), facecolors=('tab:red'))
            z += 1
    name = 'chart.png'
    plt.tight_layout()
    plt.savefig(name, dpi=799)
    plt.close()
    return name


async def get_data_chart_guntt():
    masters = []
    findata = []
    dates = []
    for x in range(15):
        days = datetime.datetime(datetime.datetime.now().year, datetime.datetime.now().month,
                                 datetime.datetime.now().day) + datetime.timedelta(days=x)
        dates.append(days.strftime('%Y-%m-%d'))
    path = os.getcwd()
    try:
        file = path + "/Config.xlsx"
    except OSError:
        file = path + "\Config.xlsx"
    wb_obj = load_workbook(filename=file)
    wsheet = wb_obj['Masters']
    for num, name in wsheet.iter_rows(values_only=True):
        if num == "#":
            continue
        masters.append(name)
    try:
        file = path + "/Schedule.xlsx"
    except OSError:
        file = path + "\Schedule.xlsx"
    wb_obj = load_workbook(filename=file, data_only=True)
    wsheet = wb_obj['Schedule']
    for mas in masters:
        temp = []
        data = []
        for date in dates:
            for num, master, from_c, to, id, phone, name, order, order_time, status in wsheet.iter_rows(values_only=True):
                if date == str(from_c)[:10] and master == mas:
                    tupl = (float(datetime.datetime.strftime(from_c, "%H.%M")),
                            float(datetime.datetime.strftime(to, "%H.%M")) -
                            float(datetime.datetime.strftime(from_c, "%H.%M")))
                    temp.append(tupl)
            data.append(temp)
            temp = []
        findata.append(data)
    result = await make_chart_guntt(findata, masters, dates)
    return result



